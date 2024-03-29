#######################################################################################################################
# Name:         INITIALIZATION - GENERAL.
# Date:         April 2023
# Description:  These functions support Jupyter Notebook 'ames-housing-pieter.ipynb' (JADS PE Foundation),
#               among many others projects.
# type: ignore
#######################################################################################################################

# The term '# type: ignore' above turns off error messages due to unknown variables even though they get assigned.
# https://github.com/microsoft/pylance-release/issues/929


#######################################################################################################################
# SYSTEM MODULES
#######################################################################################################################

# Load packages as a whole and as-is.
import re
import os
import time

# Load packages as a whole under an alias.
import pandas as pd
import numpy as np

# Load selected functions from packages.
from datetime import datetime


#######################################################################################################################
# PARAMETERS
#######################################################################################################################

# Path to code folder of this project.
C_PATH_CODE = os.getcwd()

# Paths to project root folder and sub-folders.
C_PATH_PROJECT_ROOT = re.sub("Code", "", C_PATH_CODE)
C_PATH_DATA         = os.path.join(C_PATH_PROJECT_ROOT, "Data")
C_PATH_DELIVERABLES = os.path.join(C_PATH_PROJECT_ROOT, "Deliverables")
C_PATH_DOCUMENTS    = os.path.join(C_PATH_PROJECT_ROOT, "Documents")
C_PATH_IMAGES       = os.path.join(C_PATH_PROJECT_ROOT, "Images")


#######################################################################################################################
# DEVELOPED FUNCTIONS
#######################################################################################################################

def f_who_am_i():

    """
    Get name and root folder of this computer.

    Parameters
    ----------
    -

    Returns
    -------
    str
        Computer name.
    str
        Root folder.
    """

    # Machine name of computer.
    C_MACHINE_NAME = os.uname().nodename
    

    # Computer name.
    if C_MACHINE_NAME in ['Pieters-Mac-Studio.local']:
        C_COMPUTER_NAME = 'macstudio'
    
    elif C_MACHINE_NAME in ['Pieters-MacBook-Pro.local', 'Pieters-MBP', 'pieters-mbp.home']:
        C_COMPUTER_NAME = 'macbookpro'

    else:
        raise ValueError('Unknown machine name, cannot determine C_COMPUTER_NAME')
        
        
    # Root folder (Partner folder).
    C_ROOT = re.search(r'.+/Partners/', os.getcwd()).group()
    
    return C_COMPUTER_NAME, C_ROOT


#######################################################################################################################


def f_info(

    x,
    n_top   = 10,
    n_width = 29
):

    """
    Get frequency information on column in data frame.

    Parameters
    ----------
    x: Pandas Series
        Column in data frame you want to analyse.
    n_top: int / str
        Maximum number of items to show. In case you want to see all items, enter 'all'.
    n_width: int
        Maximum number of characters to show of the values. This is useful in case the values consist of (long) sentences.

    Returns
    -------
    -
        Printed output.

    Testing
    -------
    x = [4, 4, 4, 5, 5, 6, 7, 7, 7, 7, np.nan]
    x = ["abcdef", "abcdef", "abcdef", "abcdefghi", "abcdefghi", "abcdefghi", "abcdefghi", ""]
    x = pd.Series(x)

    df_ames = pd.read_csv('https://raw.githubusercontent.com/jads-nl/discover-projects/main/ames-housing/AmesHousing.csv')
    x = df_ames['Pool QC']
    x = df_ames['Lot Frontage']
    x = df_files_subset['file_name']

    x = df_data.competentie
    n_top   = 50
    n_width = 50
    """


#----------------------------------------------------------------------------------------------------------------------
# ERROR CHECK
#----------------------------------------------------------------------------------------------------------------------

    if(not isinstance(x, pd.Series) and not isinstance(x, list)):
        raise TypeError(f"You provided an invalid type for 'x'; it must be a pandas series or a list.")


    if(not isinstance(n_top, int) and n_top != "all"):
        raise TypeError(f"You provided an invalid type for 'n_top' ('{n_top}'); it must be 'all' or an integer.")


#----------------------------------------------------------------------------------------------------------------------
# INITIALIZATION
#----------------------------------------------------------------------------------------------------------------------

    if(isinstance(x, list)):
        l_input = pd.Series(x.copy())
    else:
        l_input = x.copy()

    # Number of elements.
    n_len = len(l_input)

    # Number of unique elements.
    n_unique = len(set(l_input))

    # Number to show.
    if(n_top == "all"):        
        n_top = n_unique

    # We take max of length and 3 to prevent count errors below. Width is at least 3.
    n_char_count = max(3, len(f"{n_len:,}"))


#----------------------------------------------------------------------------------------------------------------------
# MAIN
#----------------------------------------------------------------------------------------------------------------------

    # Calculate basic info.
    df_basic_info = pd.DataFrame({

        'x': [

            "Total elements:",
            "Unique elements:",
            "empty:",
            "pd.isna():"
        ],

        'y': [                

            f"{len(l_input):,}".rjust(n_char_count),
            f"{n_unique:,}".rjust(n_char_count),
            f"{pd.Series(l_input=='').sum():,}".rjust(n_char_count),
            f"{sum(pd.isna(x) for x in l_input):,}".rjust(n_char_count)
        ],

        'z': [
            
            "",
            "",
            f"{round(pd.Series(l_input=='').sum() / n_len * 100, 1)}%".rjust(4),
            f"{round(sum(pd.isna(x) for x in l_input) / n_len * 100, 1)}%".rjust(4)
        ]
    })

    # Append numerical statistics in case x contains numerical data.
    if isinstance(l_input.values[0], (int, float, complex)):

        df_basic_info = pd.concat(

            [
                df_basic_info,
            
                pd.DataFrame({

                    'x': [

                        "0:",
                        "Inf(-):",
                        "Inf(+):"
                    ],

                    'y': [                

                        f"{sum(l_input==0):,}".rjust(n_char_count),
                        f"{sum((x is float('-inf') for x in l_input)):,}".rjust(n_char_count),
                        f"{sum((x is float('inf')  for x in l_input)):,}".rjust(n_char_count)
                    ],

                    'z': [

                        f"{round(sum(l_input==0) / n_len * 100, 1)}%".rjust(4),
                        f"{round(sum((x == float('-inf') for x in l_input)) / n_len * 100, 1)}%".rjust(4),
                        f"{round(sum((x == float('inf') for x in l_input)) / n_len * 100, 1)}%".rjust(4)
                    ]
                })
            ]
        )

    # Show in console, left align.
    c_0 = f"{0}".rjust(n_char_count)

    df_basic_info = df_basic_info.query("y != @c_0")
    df_basic_info.columns = ["="*(n_width-1), "="*n_char_count, "="*5]
    df_basic_info.index = [' ']*len(df_basic_info)

    # Replace any NaN and/or None by "None".
    l_input = l_input.fillna("NA")
    l_input = l_input.replace(float('-inf'), "-Inf ")
    l_input = l_input.replace(float('inf'), "Inf ")

    # Frequency table
    ps_freq = l_input.value_counts()

    # Calculate frequency of levels in vector.
    df_freq_source = pd.DataFrame({

        'value': ps_freq.index,
        'freq':  ps_freq.values
    })

    # Sort.
    df_freq_source = df_freq_source.sort_values(by=['freq', 'value'], ascending=[False, True])

    # Reduce length if len(x) > n_width.
    df_freq_source.value = [

        str(x)[0:(n_width)] + "..." if len(str(x)) >= (n_width - 0) else str(x)
        
        for x in df_freq_source.value # x = df_freq_source.value[0]
    ]

    # Define df_freq.
    df_freq_source['freq2'] = [f"{x}".rjust(n_char_count) for x in df_freq_source.freq]
    df_freq_source['perc']  =  df_freq_source.freq / sum(df_freq_source.freq) * 100
    df_freq_source['perc2'] = [f"{round(x,1)}%".rjust(4) for x in df_freq_source.perc]

    # Define df_dots.
    df_dots = pd.DataFrame({
        
        'value': "...",
        'freq': " "*(n_char_count - 3) + "...",
        'perc': " "*2                  + "..."
        },
        index = [0]
    )

    # Define df_total.
    df_total = pd.DataFrame({

            'value': ["-"*(n_width-1),  "TOTAL"],
            'freq':  ["-"*n_char_count, f"{n_len:,}"],
            'perc':  ["-"*5,            " 100%"]
    })

    # Update frequency section.
    df_freq = df_freq_source.drop(['freq', 'perc'], axis=1)
    df_freq = df_freq.rename(columns={'freq2':'freq', 'perc2':'perc'})
    df_freq = df_freq.head(n_top)

    # Puntjes toevoegen als n.top een getal is.
    if isinstance(n_top, int) and n_top < n_unique:
        df_freq = pd.concat([df_freq, df_dots])

    # Total toevoegen.
    df_freq         = pd.concat([df_freq, df_total])
    df_freq.columns = df_basic_info.columns
    df_freq.index   = ['']*len(df_freq)

    # Table strings.
    #c_type_table = "Type: " + type(l_input[0]).__name__

    c_freq_table = f_ifelse(
        
        isinstance(n_top, int),

        f_ifelse(

            n_unique <= n_top,
            "All items:",
            "Top-" + str(n_top)
        ),

        "All items:"

        ) + " (type: '" + type(l_input.values[0]).__name__ + "')"

    # Header frequency table.
    
    print("\n  " + " "*(n_width + n_char_count) + "n  perc")
    print(df_basic_info)
    print("\n  " + c_freq_table + " "*(n_width + n_char_count - len(c_freq_table)) + "n  perc")
    print(df_freq)


#######################################################################################################################

def f_join(l_input, c_sep = ',', b_quote = False, c_quote = "'", c_and = None):

    """
    Join list of items separated by ','.

    Parameters
    ----------
    l_input     : list
        List of items to collapse.
    c_sep:      : str
        Separator (default: ',').
    b_quote     : boolean
        Should items be quoted? (default: False).
    c_quote     : str
        The quote to be used, typically '"' or "'".
    c_and       : str
        Binding element between (for-)last elements (default: None).

    Returns
    -------
    str
        The joined string.

    Testing
    -------

    # In alle gevallen:
    c_sep   = ','
    b_quote = False
    c_quote = "'"
    c_and   = None

    l_input  = ['apple', 'banana', 'pear', 5]
    c_sep    = ','
    b_quote  = True
    c_quote  = "'"
    c_and    = 'and'

    f_join(l_input)

    l_input = df_non_allowed_score.competentie
    c_and   = 'en'
    b_quote = True

    """  

    ###################################################################################################################
    # Initialization.
    ###################################################################################################################

    # Check whether l_input is a list, if not make it a list.
    if isinstance(l_input, pd.Series):
            l_input = list(l_input)

    # Determine length of l_input.
    n_length = len(l_input)

    # Add quotation if requested.
    if b_quote:

        l_input = [c_quote + item + c_quote for item in map(str, l_input)]

    # Add space behind c_sep if it is not "\n", e.g., when it is ","".
    c_sep = c_sep if c_sep == "\n" else c_sep + " "


    ###################################################################################################################
    # Main.
    ###################################################################################################################

    if c_and is None:

        c_output = c_sep.join(l_input)

    else:

        if n_length == 1:
        
            c_output = l_input[0]

        elif n_length == 2:

            c_output = l_input[0] + " " + c_and + " " + l_input[1]

        else:

            c_output = c_sep.join(l_input[:-1]) + c_sep + c_and + " " + l_input[n_length-1]


    ###################################################################################################################
    # Return.
    ###################################################################################################################

    return c_output


#######################################################################################################################

def f_grepl(pattern, l_str):

    """
    Searches for matches to argument 'pattern' within each element of a character list, 'l_str'.
    The Python equivalent of 'grepl' in R.

    Parameters
    ----------
    pattern : 'str'
        Regex pattern.
    l_str : 'list'
        Character list.

    Returns
    -------
    list
        Boolean list, True in case of a match and False in case of a non-match.

    Testing
    -------
    f_grepl("^P", ["Pieter", "Bart", "Theo", "aPieter"])
    """

    return [bool(re.search(pattern, x)) for x in l_str]


#######################################################################################################################

# Ifelse
def f_ifelse(b_eval, true, false):

    """
    <short description>.

    Parameters
    ----------
    <name> : <type>
        <short description>.
    <name> : <type>
        <short description>.

    Returns
    -------
    <type>
        <short description>.
    """

    if b_eval:
        output = true
    else:
        output = false

    return output


#######################################################################################################################

# Test whether number is number
def f_is_numerical(value):

    try:
        float(value)
        return True
    
    except ValueError:
        return False
        

#######################################################################################################################

def f_clean_up_header_names(l_input):

    """
    Clean up header names of data frame: (1) set names to lower case, and (2) replace spaces by '_'.

    Parameters
    ----------
    l_input : list
        Column names.


    Returns
    -------
    list
        Cleaned up column names.
    """
    

    return [
        # Put in lower case:
        x3.lower() for x3 in [

        # Replace space by '_':
        x2 if f_is_numerical(x2) else re.sub(" |\.", "_", x2) for x2 in [

        str(x1) for x1 in l_input        
    ]]]


#######################################################################################################################

def f_check_nonnumeric_in_df(
    
    df_input,
    l_exclude_columns = [],
    l_include_columns = []
    ):

    """
    Check on non-numeric in data frame.
    """

    """
    <short description>.

    Parameters
    ----------
    <name> : <type>
        <short description>.
    <name> : <type>
        <short description>.

    Returns
    -------
    <type>
        <short description>.

    Testing
    -------
    df_input          = df_nline_w_source
    l_exclude_columns = ['product']
    l_include_columns = []
    """

        
    # Error check - Are all column names in 'l_exclude_column' and 'l_include_column' present in df_input?
    l_exclude_columns_not_in_df_input = [x for x in l_exclude_columns if x not in  df_input.columns]
    l_include_columns_not_in_df_input = [x for x in l_include_columns if x not in  df_input.columns]

    if len(l_exclude_columns_not_in_df_input) > 0:

        c_temp = ', '.join(f"'{x}'" for x in l_exclude_columns_not_in_df_input)
        
        raise ValueError(
            f"The following column name(s) in 'l_exclude_columns' are not present in the column names of 'df_input': {c_temp}"
        )

    if len(l_include_columns_not_in_df_input) > 0:

        c_temp = ', '.join(f"'{x}'" for x in l_include_columns_not_in_df_input)
        
        raise ValueError(
            f"The following column name(s) in 'l_include_columns' are not present in the column names of 'df_input': {c_temp}"
        )


    # Error check - Are the same column names present in both 'l_exclude_column' and 'l_include_column'?
    l_overlap_include_exclude_columns = set(l_include_columns).intersection(set(l_exclude_columns))

    if len(l_overlap_include_exclude_columns) > 0:

        c_temp = ', '.join(f"'{x}'" for x in l_overlap_include_exclude_columns)
        
        raise ValueError(
            f"The following column name(s) are present in both 'l_exclude_columns' and 'l_include_columns': {c_temp}"
        )


    # Initialization.
    if l_include_columns == []:
        l_include_columns = df_input.columns


    # Main.
    df_to_check = df_input[[x for x in df_input.columns if x in l_include_columns and x not in l_exclude_columns]]
    df_eval     = df_to_check[~df_to_check.applymap(np.isreal).all(axis = 1)]

    if df_eval.shape[0] > 0:

        print(
            f"\nWARNING - '{f_var_name(df_input)}' contains non-numerical. We observe {df_eval.shape[0]} row(s) with at "
            f"least one non-numerical, below we show the first 5 rows (at max):\n"
        )

        print(df_eval.head(5))

        print("\nFor reference, the full data frame, incl. those columns that were not evaluated:\n")

        print(df_input.filter(items = df_eval.index, axis=0).head(5))

        print("\n")

    else:

        print(f"\nOK - '{f_var_name(df_input)}' contains numericals only.\n")


#######################################################################################################################

def f_var_name(var):

    """
    Get argument name of variable assigned to function parameter. It is used in 'f_describe()' for example.

    Parameters
    ----------
    var : any
        Parameter name to which an object is assigned to, and of which we want the name.

    Returns
    -------
    str
        Name of object assigned to parameter, i.e., the argument name.
    """

    lcls = inspect.stack()[2][0].f_locals

    for name in lcls:        
        if id(var) == id(lcls[name]):
            return name

    return None

#######################################################################################################################


def f_check_na_in_df(
    
    df_input,
    l_exclude_columns = [],
    l_include_columns = []
    ):

    """
    Check on empty cells in data frame.

    Parameters
    ----------
    <name> : <type>
        <short description>.
    <name> : <type>
        <short description>.

    Returns
    -------
    <type>
        <short description>.

    Testing
    -------
    df_input          = df_htri_w_source
    l_exclude_columns = ['product']
    l_include_columns = []
    """

            
    # Error check - Are all column names in 'l_exclude_column' and 'l_include_column' present in df_input?
    l_exclude_columns_not_in_df_input = [x for x in l_exclude_columns if x not in  df_input.columns]
    l_include_columns_not_in_df_input = [x for x in l_include_columns if x not in  df_input.columns]

    if len(l_exclude_columns_not_in_df_input) > 0:

        c_temp = ', '.join(f"'{x}'" for x in l_exclude_columns_not_in_df_input)
        
        raise ValueError(
            f"The following column name(s) in 'l_exclude_columns' are not present in the column names of 'df_input': {c_temp}"
        )

    if len(l_include_columns_not_in_df_input) > 0:

        c_temp = ', '.join(f"'{x}'" for x in l_include_columns_not_in_df_input)
        
        raise ValueError(
            f"The following column name(s) in 'l_include_columns' are not present in the column names of 'df_input': {c_temp}"
        )


    # Error check - Are the same column names present in both 'l_exclude_column' and 'l_include_column'?
    l_overlap_include_exclude_columns = set(l_include_columns).intersection(set(l_exclude_columns))

    if len(l_overlap_include_exclude_columns) > 0:

        c_temp = ', '.join(f"'{x}'" for x in l_overlap_include_exclude_columns)
        
        raise ValueError(
            f"The following column name(s) are present in both 'l_exclude_columns' and 'l_include_columns': {c_temp}"
        )


    # Initialization.
    if l_include_columns == []:
        l_include_columns = df_input.columns


    # Main.
    df_to_check = df_input[[x for x in df_input.columns if x in l_include_columns and x not in l_exclude_columns]]
    df_eval     = df_to_check[df_to_check.applymap(pd.isnull).any(axis = 1)]


    if df_eval.shape[0] > 0:

        print(
            f"\nWARNING - '{f_var_name(df_input)}' contains NA. We observe {df_eval.shape[0]} row(s) with at "
            f"least one NA, below we show the first 5 rows (at max):\n"
        )

        print(df_eval.head(5))

        print(
            f"\nFor reference, the full data frame, incl. those columns that were not evaluated. "
            "Below, we show at max the first 5 rows:\n"
        )

        #print(df_input.filter(items = df_eval.index, axis=0).head(5))
        print(df_input.head(5))

        print("\n")

    else:

        print(f"\nOK - '{f_var_name(df_input)}' is fully filled.\n")


#######################################################################################################################

def f_get_latest_file(

    c_name,
    c_path,
    c_type
    ):

    """
    Get latest file with said string in the file residing in said path.

    Parameters
    ----------   
    c_name: 'str'
        String in the file name.
    c_path: 'str'
        Path where file resides.
    c_type: 'str'
        Reference to file type to be read.

    Returns
    -------
    Pandas Series
        file: file name
        date_mod: modification date
        age: age of file as string

    Testing
    -------  
    c_name = 'HTRI'
    c_path = C_PATH_DELIVERABLES
    c_type = 'xlsx'

    c_name = 'Den Haag - Displays - Overzicht na plaatsing'
    c_path = c_path_file_xls
    c_type = 'xlsx'

    f_get_latest_file(c_name, c_path, c_type)
    """ 


#----------------------------------------------------------------------------------------------------------------------
# Initialization.
#----------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------
# Error check.
#----------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------
# Main.
#----------------------------------------------------------------------------------------------------------------------

    # Get all files in said folder, excl. any folders. I replaced f_find_str by list(filter(re.compile(c_name).search, list))
    df_file = pd.DataFrame({
        'file': list(filter(
            
            # String to search for in the file names.
            re.compile(c_name).search,

            # List with all files in c_path.
            [                
                f for f in os.listdir(c_path)

                # Filter on files only (excl dirs) and on the requested file type.
                if os.path.isfile(os.path.join(c_path, f)) and
                    os.path.splitext(os.path.join(c_path, f))[1]== '.'+c_type
            ]
        ))
    })

    # Error check - Is a file found?
    if df_file.shape[0] == 0:
        raise LookupError(
            f"No file found for:\nFile name: '{c_name}'\nFile type: '{c_type}'\nFile path: '{c_path}'"
        )

    # Add number of seconds since epoch.
    df_file.insert(1, 'date_mod_sec',
        [os.path.getmtime(os.path.join(c_path, f)) for f in df_file.file]
    )
    
    # Get first row (latest file).
    ps_file = (df_file
        .sort_values(by='date_mod_sec', ascending=False)
        .iloc[0]
    )

    # Convert seconds to time stamp.
    ps_file['date_mod'] = datetime.fromtimestamp(
        
        ps_file.date_mod_sec
        
        ).strftime('%Y-%m-%d %H:%M:%S')
  
    # Add age of file
    n_age = time.time() - ps_file.date_mod_sec

    if n_age < 60:
        c_age = 'sec'
    elif n_age < 3600:
        n_age = n_age / 60
        c_age = 'minutes'
    elif n_age < 3600*24:
        n_age = n_age / 3600
        c_age = 'hours'
    else:
        n_age = n_age / 3600 / 24
        c_age = 'days'

    ps_file['age'] = str(round(n_age,1)) + " " + c_age


#----------------------------------------------------------------------------------------------------------------------
# Return results.
#----------------------------------------------------------------------------------------------------------------------

    # Return pandas series with information, except 'date_mod_sec' (not needed).
    return ps_file.drop('date_mod_sec')


#######################################################################################################################

def f_read_data_from_file(

    l_name,
    c_path,
    c_type               = 'xlsx',
    c_sheet              = None,
    c_sep                = ',',
    l_cols               = None,
    n_skiprows           = None,
    n_rows               = None,
    n_header             = 0,
    b_clean_header_names = True,
    b_strip_spaces       = True
    ):

    """
    Read data from file into a data frame object.

    Parameters
    ----------   
    l_name: 'str'
        List of (parts of) of file names where data is to be read from.
    c_path: 'str'
        Path where file resides.
    c_type: 'str'
        Reference to file type to be read (default: 'xlsx').
    c_sheet: 'str'
        Sheet name in case data is to read from Excel file (default: 'None').
    l_cols: 'int',
        If list of int, then indicates list of column numbers to be parsed (0-indexed).
    n_skiprows: 'int'
        Line numbers to skip (0-indexed) or number of lines to skip (int) at the start of the file. 
    n_rows: 'int'
        Number of rows to parse (default: None).
    n_header: 'int'
        Row (0-indexed) to use for the column labels of the parsed DataFrame.
    b_clean_header_names: 'bool'
        Do we clean up the header names? (default: True)
    b_strip_spaces: 'bool'
        Do we strip spaces before and after the data in each cell? (default: False)

    Returns
    -------
    Data frame
        Data read from file is stored in data frame object.

    Testing
    -------  
    Default:
    c_type               = 'xlsx'
    c_sheet              = None
    c_sep                = ','
    l_cols               = None
    n_skiprows           = None
    n_rows               = None
    n_header             = 0
    b_clean_header_names = True
    b_strip_spaces       = True


    l_name  = ["Classification GIs - " + x for x in l_comp_list]
    c_path  = C_PATH_DATA
    

    l_name  = 'Content Database'
    c_path  = C_PATH_DATA
    c_type  = 'xlsx'
    c_sheet = 'gedragsindicatoren'

    """ 


#----------------------------------------------------------------------------------------------------------------------
# Initialization.
#----------------------------------------------------------------------------------------------------------------------

    # Valid file types.
    l_type_valid = ['xlsx', 'xlsm', 'csv', 'parquet']

    # Convert l_name to list - with string as single element - in case it is a string.
    if isinstance(l_name, str):
        l_name = [l_name]

    # Latest file per file name. c_name='Content Database'
    l_file = [

        f_get_latest_file(c_name, c_path, c_type)

        for c_name in l_name
    ]

    # Create empty list
    l_df_data = []


#----------------------------------------------------------------------------------------------------------------------
# Error check.
#----------------------------------------------------------------------------------------------------------------------

    if c_type not in l_type_valid :
        raise ValueError(f"You did not provide a valid file type. Choose 'c_type' to be one of {', '.join(l_type_valid)}.")


#----------------------------------------------------------------------------------------------------------------------
# Main.
#----------------------------------------------------------------------------------------------------------------------

    # Iterate through all file names.
    for i in range(len(l_name)): # i=0

        # Excel
        if c_type in ['xlsx', 'xlsm']:            
                
            obj = pd.read_excel(

                io         = os.path.join(c_path, l_file[i].file),
                sheet_name = c_sheet,
                usecols    = l_cols,
                skiprows   = n_skiprows,
                nrows      = n_rows,
                header     = n_header,
                engine     = 'openpyxl'
            )

            # If we don't supply a sheet name the output is a dictionary of data frames,
            # from which we will take the first worksheet.
            if isinstance(obj, dict):
                    
                c_sheet_temp = list(obj.keys())[0]
                
                df_temp = obj[c_sheet_temp]

            # When sheet name does exist the output is a data frame.
            else:

                c_sheet_temp = None

                df_temp = obj


            # Append data frame to list of data frames.
            l_df_data.append(df_temp)


        # CSV
        if c_type == 'csv':

            l_df_data.append(
                
                pd.read_csv(

                    filepath_or_buffer = os.path.join(c_path, l_file[i].file),
                    sep                = c_sep,
                    usecols            = l_cols,
                    skiprows           = n_skiprows,
                    header             = n_header
                )
            )


        # Parquet
        if c_type == 'parquet':

                l_df_data.append(

                    pd.read_parquet(

                        path    = os.path.join(c_path, l_file[i].file),
                        engine  = 'pyarrow',
                        columns = l_cols
                    )
                )

       

        # Comms to the user.
        print(f"\nReading at : {datetime.now()}")

        print(f"Requested  : '{l_name[i]}' (file), '{c_type}' (type)")

        print(f"Read file  : '{l_file[i].file}'")

        if c_type in ['xlsx', 'xlsm']:
            
            if c_sheet is not None:
                
                print(f"Sheet name : '{c_sheet}' - provided by you.")

            else:

                print(f"Sheet name : '{c_sheet_temp}' - first sheet in the workbook.")


        print(f"Path       : '.../{re.sub(f_who_am_i()[1], '', c_path)}'")

        print(f"Modified   : {l_file[i].date_mod}")

        print(f"Age        : {l_file[i].age}")

        print(f"==========================")


    # Concatenate data frames.
    df_data = pd.concat(l_df_data)

    # Clean up header names?
    if b_clean_header_names:
        df_data.columns = f_clean_up_header_names(l_input = df_data.columns)


    # Strip spaces before and after the data in each cell?
    if b_strip_spaces:
        df_data = df_data.map(lambda x: x.strip() if isinstance(x, str) else x)


#----------------------------------------------------------------------------------------------------------------------
# Return results.
#----------------------------------------------------------------------------------------------------------------------

    return df_data

#######################################################################################################################

def f_now():

    """
    Get string containing today's date and current time.

    Parameters
    ----------
    -

    Returns
    -------
    str
        String containing today's date and current time.

    Testing
    -------
    """ 

    # Current time.
    dt_now = datetime.now()

    return(
        re.sub("-", " ", str(dt_now.date())) + " - " +
        dt_now.strftime("%H %M %S")
    )


#######################################################################################################################

def f_write_data_to_file(

    l_df,
    c_name = 'Temp',
    c_path = C_PATH_DATA,
    c_type = 'xlsx',
    l_name = None
    ):

    """
    Write object to file.

    Parameters
    ----------
    l_df: 'list' or 'Pandas Series' of values, 'Pandas DataFrame', or 'list' of 'Pandas DataFrame's.
        Data object to write to file.    
    c_name: 'str'
        Name of the file where data object will be saved in.
    c_path: 'str'
        Path where file will be saved.
    c_type: 'str'
        Reference to file type (default: 'xlsx').
    l_name: 'str', or 'list' of 'str'
        Names to be used as sheet names or added to file name (default: 'None').

    Returns
    -------
    -
        Print statement in console to confirm writing of data to file.

    Testing
    -------
    l_df          = [pd.DataFrame({'a': [1,2,3,2,3,3], 'b': [5,6,7,8,9,9]}), pd.DataFrame({'a': [1,2,3], 'b': [5,6,7]})]
    l_df          =  pd.DataFrame({'a': [1,2,3,2,3,3], 'b': [5,6,7,8,9,9]})
    l_df          = pd.Series([1,2,3,4])
    l_df          = [1,2,3,2,3,3]    
    c_name = "Data file"
    c_path        = C_PATH_DELIVERABLES
    c_type        = 'xlsx'
    c_type        = 'csv'
    c_type        = 'parquet'
    l_name        = None
    l_name        = ['DATA1', 'DATA2']

    f_write_data_to_file(l_df, c_name, c_path, c_type, l_name)
    """ 


#----------------------------------------------------------------------------------------------------------------------
# Initialization.
#----------------------------------------------------------------------------------------------------------------------

    # Assign object name, for later use when communicating to user, see end.
    c_l_df = f_var_name(l_df)

    # Valid file types.
    l_type_valid = ['xlsx', 'csv', 'parquet']

    # Current date and time.
    c_now  = f_now() + " - "


    # Check on type of l_df and make corrections as needed.
    if isinstance(l_df, list) and not isinstance(l_df[0], pd.DataFrame):
        l_df = pd.Series(l_df)

    if isinstance(l_df, pd.Series):
        l_df = pd.DataFrame({'l_df': l_df})

    if isinstance(l_df, pd.DataFrame):
        l_df = [l_df]


    # Convert l_name to list if it is a str.
    if isinstance(l_name, str):
        l_name = [l_name]

    # Create l_name if not provided.
    if l_name is None:
        l_name = ['data' + str(i+1) for i in f_seq_along(l_df)]


#----------------------------------------------------------------------------------------------------------------------
# Error check.
#----------------------------------------------------------------------------------------------------------------------

    if len(l_df) != len(l_name):
        raise IndexError("Length of 'l_df' and 'l_name' are not the same.")

    if c_type not in l_type_valid :
        raise ValueError(f"You did not provide a valid file type. Choose 'c_type' to be one of {', '.join(l_type_valid)}.")


#----------------------------------------------------------------------------------------------------------------------
# Main.
#----------------------------------------------------------------------------------------------------------------------

    # Excel - Store dataframe(s) in separate worksheets in same workbook.
    # To check later - https://xlsxwriter.readthedocs.io/example_pandas_table.html
    if c_type == 'xlsx':
       
        from openpyxl import load_workbook
        from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment

        # with pd.ExcelWriter(os.path.join(c_path, c_now + c_name + "." + c_type)) as writer:

        #     for i in range(len(l_df)):

        #         l_df[i].to_excel(
        #             excel_writer = writer,
        #             sheet_name   = l_name[i],
        #             index        = False
        #         )

        with pd.ExcelWriter(
            
            path   = os.path.join(c_path, c_now + c_name + "." + c_type),
            engine = 'openpyxl',
            mode   = 'w'

        ) as writer:
            
            # Access the workbook
            #workbook = writer.book

            for i in range(len(l_df)):

                l_df[i].to_excel(
                    excel_writer = writer,
                    sheet_name   = l_name[i],
                    index        = False,
                    engine       = 'openpyxl'
                )

                # Access the worksheet.
                worksheet = writer.sheets[l_name[i]]

                # Set the zoom level for the worksheet to 150%.
                worksheet.sheet_view.zoomScale = 150

                # # Set column width for columns C and D.
                # dc_col_width = {'A': 15, 'B': 10, 'C': 50, 'D': 50, 'E': 10, 'F': 10, 'G': 50}

                # for k, v in dc_col_width.items():

                #     # Set the column width for column C and D to width of 85.
                #     worksheet.column_dimensions[f'{k}'].width = v

                #     # # Wrap text in columns.
                #     # for cell in worksheet[f'{col}']:
                #     #     cell.alignment = Alignment(wrapText=True)

                # # Align text in vertical direction.
                # for col in ['A', 'B', 'C', 'D', 'E']:

                #     # Wrap and vertically center text in all columns.
                #     for cell in worksheet[f'{col}']:

                #         cell.alignment = Alignment(
                #             vertical = 'center',
                #             wrapText = True
                #         )

                # # Define the name of the style you want to create or modify
                # style_name = 'TableStyle'

                # # Check if the style already exists in the workbook
                # style_found = False
                # for named_style in workbook.named_styles:
                #     if named_style.name == style_name:

                #         # If the style exists, modify it
                #         table_style = named_style
                #         style_found = True
                #         break

                # if not style_found:
                #     # If the style does not exist, create a new named style
                #     table_style = NamedStyle(name=style_name)

                # # Apply font style
                # table_style.font = Font(
                #     #bold = True,
                #     name = 'Arial',  # Specify the font name (e.g., Arial)
                #     size = 11        # Specify the font size
                # )

                # # Apply fill color.
                # table_style.fill = PatternFill(
                #     start_color = 'FFFF00',
                #     end_color   = 'FFFF00',
                #     fill_type   = 'solid'
                # )

                # # Apply borders
                # border = Border(
                #     left=Side(border_style='thin', color='000000'),
                #     right=Side(border_style='thin', color='000000'),
                #     top=Side(border_style='thin', color='000000'),
                #     bottom=Side(border_style='thin', color='000000')
                # )
                # table_style.border = border

                # # Apply text alignment.
                # table_style.alignment = Alignment(
                #     #horizontal = 'center',
                #     vertical   = 'center',
                #     wrapText   = True
                # )

                # # Apply the style to a range (e.g., the entire table).
                # for row in worksheet.iter_rows(                    
                #     min_row = 2,
                #     max_row = len(l_df[i]) + 1,
                #     min_col = 1,
                #     max_col = len(l_df[i].columns)
                # ):
                #     for cell in row:
                #         cell.style = table_style

                # Save the Excel file
                #writer.save()







                ## I NEED TO MAKE STYLE CHANGES AFTER I SAVED IT
                ## THE SCRIPT BELOW I GOT FROM CHAT GPT

                # pandas.ExcelWriter in the 'openpyxl' engine doesn't provide direct access to named styles to update or modify them.
                # To update styles using pandas.ExcelWriter, you'll typically need to rely on the native Excel styling capabilities 
                # available through openpyxl or by manually modifying the Excel file after it has been generated by pandas.

                # Here's an example of how to update styles after writing an Excel file using pandas.ExcelWriter:

                # import pandas as pd
                # from openpyxl import load_workbook
                # from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                # # Create a sample DataFrame
                # data = {'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']}
                # df = pd.DataFrame(data)

                # # Create an ExcelWriter object using openpyxl engine
                # with pd.ExcelWriter('styled_table.xlsx', engine='openpyxl', mode='w') as writer:
                #     df.to_excel(writer, sheet_name='Sheet1', index=False)

                # # Load the generated Excel file
                # workbook = load_workbook('styled_table.xlsx')

                # # Access a specific sheet within the workbook (replace 'Sheet1' with your sheet name)
                # sheet = workbook['Sheet1']

                # # Access the specific worksheet
                # worksheet = sheet

                # # Define a Font object for font formatting
                # font = Font(
                #     bold=True,
                #     name='Arial',  # Specify the font name (e.g., Arial)
                #     size=12,       # Specify the font size
                # )

                # # Apply the font formatting to a range of cells (e.g., A2:B4)
                # for row in worksheet.iter_rows(
                #     min_row=2,
                #     max_row=4,
                #     min_col=1,
                #     max_col=2
                # ):
                #     for cell in row:
                #         cell.font = font

                # # Define other styles (fill color, borders, alignment, etc.) and apply them as needed

                # # Save the modified workbook
                # workbook.save('styled_table_updated.xlsx')







    # CSV - Store dataframe(s) in separate CSV files.
    if c_type == 'csv':

        for i in range(len(l_df)):

            l_df[i].to_csv(
                path  = os.path.join(c_path, c_now + c_name + " - " + l_name[i] + "." + c_type),
                index = False
            )


    # Parquet - Store dataframe(s) in separate CSV files.
    if c_type == 'parquet':

        for i in range(len(l_df)):

            l_df[i].to_parquet(
                path   = os.path.join(c_path, c_now + c_name + " - " + l_name[i] + "." + c_type),
                index  = False,
                engine = 'pyarrow'
            )


    # l_df[0].iloc[:,:5].to_parquet(c_path + c_now + c_name + " - " + l_name[i] + "." + c_type, index=False)

    # Comms to the user.
    print(f"\nWriting at : {datetime.now()}")

    print(f"Object     : '{c_l_df}'")

    print(f"Name       : '{c_now + c_name + '.' + c_type}'")

    print(f"As         : '{c_type}'")

    print(f"Path       : '.../{re.sub(f_who_am_i()[1], '', c_path)}'")

    print(f"==========================")

#----------------------------------------------------------------------------------------------------------------------
# Return results.
#----------------------------------------------------------------------------------------------------------------------


#######################################################################################################################

def f_seq_along(l_input):

    """
    Generate regular sequence as long as the provided list.
        
    Parameters
    ----------
    l_input : list
        List to determine length of sequence.
   
    Returns
    -------
    list
        Sequence as long as l_input.

    Test
    ----
    l_input = ['a', 'b', 'c']
    l_input = pd.Series(l_input)
    f_seq_along(l_input)
    """


#----------------------------------------------------------------------------------------------------------------------
# Initialization.
#----------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------
# Error check.
#----------------------------------------------------------------------------------------------------------------------

    if not isinstance(l_input, (list, pd.Series)):
        raise IndexError("Length of 'x' and 'l_name' are not the same.")

#----------------------------------------------------------------------------------------------------------------------
# Main.
#----------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------
# Return results.
#----------------------------------------------------------------------------------------------------------------------

    return list(np.arange(len(l_input)))


#######################################################################################################################

# def f_(

#     ):

#     """
#     <short description>.

#     Parameters
#     ----------
#     <name> : <type>
#         <short description>.
#     <name> : <type>
#         <short description>.

#     Returns
#     -------
#     <type>
#         <short description>.
#     """


# #----------------------------------------------------------------------------------------------------------------------
# # Testing.
# #----------------------------------------------------------------------------------------------------------------------

# #----------------------------------------------------------------------------------------------------------------------
# # Initialization.
# #----------------------------------------------------------------------------------------------------------------------

# #----------------------------------------------------------------------------------------------------------------------
# # Error check.
# #----------------------------------------------------------------------------------------------------------------------

# #----------------------------------------------------------------------------------------------------------------------
# # Main.
# #----------------------------------------------------------------------------------------------------------------------

# #----------------------------------------------------------------------------------------------------------------------
# # Return results.
# #----------------------------------------------------------------------------------------------------------------------

#     return

#######################################################################################################################